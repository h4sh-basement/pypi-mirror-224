"""
Data type conversion of different files
"""
import json
from openpyxl import load_workbook
from pathlib import Path

mypath = Path.cwd()


def excel_to_list(file=None, sheet="Sheet1", line=1):
    """
    Convert Excel file data to list
    :param file: Path to file
    :param sheet: excel sheet, default name is Sheet1
    :param line: Start line of read data
    :return: list data
    @data("data.xlsx", sheet="Sheet1", line=1)
    def test_login(self, username, password):
        print(username)
        print(password)
    """
    if file is None:
        raise FileExistsError("Please specify the Excel file to convert.")

    excel_table = load_workbook(file)
    sheet = excel_table[sheet]

    table_data = []
    for line in sheet.iter_rows(line, sheet.max_row):
        line_data = []
        for field in line:
            line_data.append(field.value)
        table_data.append(line_data)

    return table_data


def _check_data(list_data: list) -> list:
    """
    Checking test data format.
    :param list_data:
    :return:
    """
    if isinstance(list_data, list) is False:
        raise TypeError("The data format is not `list`.")
    if len(list_data) == 0:
        raise ValueError("The data format cannot be `[]`.")
    if isinstance(list_data[0], dict):
        test_data = []
        for data in list_data:
            line = []
            for d in data.values():
                line.append(d)
            test_data.append(line)
        return test_data
    else:
        return list_data


def json_to_list(file, key=None):
    """
    Convert JSON file data to list
    :param file: Path to file
    :param key: Specifies the key for the dictionary
    :return: list data
    @data("run.bat", key="login")
    def test_login(self, username, password):
        print(username)
        print(password)
    """
    if file is None:
        raise FileExistsError("Please specify the JSON file to convert.")

    if key is None:
        with open(file, "r", encoding="utf-8") as f:
            data = json.load(f)
            list_data = _check_data(data)
    else:
        with open(file, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)[key]
                list_data = _check_data(data)
            except KeyError:
                raise ValueError("Check the test data, no '{}'".format(key))

    return list_data


def getexlist(file=mypath / 'test_data' / 'api_Test_data.xlsx', line=2, sheet='Access_API'):
    getexcellist = excel_to_list(file, sheet, line)
    xlisn = []
    for xlista in getexcellist:
        testcase, xendpoint, xmethod, payload, exresults, scvarib, ctype, rtype = xlista
        xlisn.append(rtype)
    return xlisn
