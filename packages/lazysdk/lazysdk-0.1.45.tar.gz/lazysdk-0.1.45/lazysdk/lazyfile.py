#!/usr/bin/env python3
# coding = utf8
"""
@ Author : ZeroSeeker
@ e-mail : zeroseeker@foxmail.com
@ GitHub : https://github.com/ZeroSeeker
@ Gitee : https://gitee.com/ZeroSeeker
"""
from . import lazypath
import collections
import subprocess
import requests
import platform
import datetime
import openpyxl
import showlog
import time
import xlrd
import json
import sys
import os


if platform.system() == 'Windows':
    path_separator = '\\'
else:
    path_separator = '/'
headers_default = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:68.0) Gecko/20100101 Firefox/68.0"}


def delete(
        file,
        postfix: str = None,
        path: str = None,
):
    """
    删除文件
    """
    if path:
        # 如果指定了路径，就加上路径
        file_dir = f'{path}{path_separator}{file}'
    else:
        # 如果没指定路径，就直接使用文件名
        file_dir = file

    if postfix:
        # 如果指定了后缀名，就加上后缀名
        file_dir = f'{file}.{postfix}'
    else:
        # 如果没指定后缀名，就忽略
        pass
    os.remove(file_dir)


def get_suffix(
        file_name: str
):
    suffix = file_name.split('.')[-1]  # 获取文件后缀名
    return suffix


def file_rename(
        before_name: str,  # 原名称
        after_name: str  # 重命名名称
):
    """
    重命名文件，注意是完整文件路径
    :param before_name:
    :param after_name:
    :return:
    """
    try:
        os.rename(before_name, after_name)
        return True
    except Exception as e:
        print(e)
        print('rename file fail\r\n')
        return False


def open_folder(
        folder: str = os.path.dirname(os.path.abspath(__file__))  # 默认为当前路径
):
    """
    功能：打开路径
    """
    if platform.system() == 'Windows':
        os.startfile(folder)  # Windows上打开文件
    else:
        subprocess.check_call(['open', folder])  # 非Windows上打开文件


def download(
        url,
        filename: str = None,
        suffix_name: str = None,
        headers: dict = None,
        path: str = "download",
        proxies=None,
        size_limit: int = None,
        range_start: int = None,
        range_end: int = None,
):
    """
    实现文件下载功能，可指定url、文件名、后缀名、请求头、文件保存路径
    :param url:
    :param filename:文件名
    :param suffix_name:后缀名
    :param headers:请求头
    :param path:文件保存路径
    :param proxies:代理
    :param size_limit:尺寸限制
    :param range_start:开始位置
    :param range_end:结束位置
    :return:
    """
    if headers is None:
        headers_local = headers_default
    else:
        headers_local = headers

    if range_start is None and range_end is None:
        range_start = 0
        range_info = None
    elif range_start is not None and range_end is None:
        range_info = 'bytes=%d-' % range_start  # 从这里向后
    elif range_start is None and range_end is not None:
        range_start = 0
        range_info = 'bytes=0-%d' % range_end
    else:
        range_info = 'bytes=%d-%d' % (range_start, range_end)

    if range_info is None:
        pass
    else:
        headers_local['Range'] = range_info
    # 获取文件的基本信息
    response = requests.get(
        url=url,
        headers=headers_local,
        stream=True,
        proxies=proxies
    )
    total_length = response.headers.get('content-length')  # 文件大小
    content_type = response.headers.get('content-type')  # 文件类型
    content_disposition = response.headers.get('content-disposition')  # 文件名及类型
    filename_default = 'unknown_' + str(time.time())
    if content_disposition is not None:
        content_dispositions = content_disposition.replace(' ', '').split(';')
        for each_content_disposition in content_dispositions:
            if 'filename' in each_content_disposition:
                each_content_disposition_split = each_content_disposition.split(sep='=', maxsplit=1)  # 只拆分一次，防止有多个=影响
                filename_default_full = each_content_disposition_split[1]
                filename_default = filename_default_full[:filename_default_full.rfind('.')]  # 解析文件名
                suffix_name = filename_default_full[filename_default_full.rfind('.')+1:]  # 解析文件后缀
            else:
                pass
    else:
        pass

    if suffix_name is None:
        # 尝试自动获取文件后缀名
        suffix_name = content_type.split('/')[1]

    if filename is None:
        download_file_name = str(filename_default) + "." + str(suffix_name)
    else:
        download_file_name = str(filename) + "." + str(suffix_name)

    if path is None:
        path_local = download_file_name
    else:
        if os.path.exists(path):
            pass
        else:
            os.makedirs(path)
        path_local = path + path_separator + download_file_name

    if range_start is None:
        temp_size = 0  # 已经下载文件大小
    else:
        temp_size = range_start + 0  # 已经下载文件大小
    chunk_size = 1024  # 分割文件大小，字节B
    total_size = int(total_length)  # 文件总大小
    total_size_mb = round(total_size / (1024 * 1024), 2)  # 换算到MB的文件大小
    # 添加文件大小控制，跳过下载超大文件
    if size_limit is None:
        pass
    else:
        if total_size > size_limit:
            return
        else:
            pass

    time_start = time.time()  # 获取下载开始时间
    is_finish = False

    with open(path_local, "ab") as f:  # wb新建文件，a追加
        for chunk in response.iter_content(chunk_size=chunk_size):
            try:
                temp_time_now = time.time()  # 时间采样
                time_spend_total = temp_time_now - time_start
                if time_spend_total == 0:
                    total_speed = round((temp_size-range_start) / (1024 * 1024) / 0.001, 2)  # 计算速度：MB/s
                else:
                    total_speed = round((temp_size-range_start) / (1024 * 1024) / time_spend_total, 2)  # 计算速度：MB/s
                if not chunk:
                    if temp_size >= total_size:
                        is_finish = True
                    else:
                        is_finish = False
                    break
                else:
                    temp_size += len(chunk)
                    f.write(chunk)
                    f.flush()
                    done = int(50 * temp_size / total_size)
                    if total_speed == 0 or time_spend_total == 0:
                        time_left = 0
                    else:
                        time_left = (total_size - temp_size) / total_speed / 1024 / 1024
                    show_dict = {
                        'finish_mark': '█' * done,
                        'not_finish_mark': ' ' * (50 - done),
                        'total_size': total_size_mb,  # 换算到M
                        'total_percent': round(100 * temp_size / total_size, 4),
                        'total_speed': total_speed,
                        'finish_size': round(temp_size / (1024 * 1024), 2),
                        'time_spend_total': int(time_spend_total),
                        'time_left': int(time_left)
                    }
                    show_msg = "\r[%(finish_mark)s%(not_finish_mark)s] " \
                               "总大小:%(total_size)sMB " \
                               "总进度:%(total_percent)s%% " \
                               "平均速度:%(total_speed)sMB/s " \
                               "已下载:%(finish_size)sMB " \
                               "已耗时 %(time_spend_total)s 秒 " \
                               "预计剩余 %(time_left)s 秒" % show_dict
                    sys.stdout.write(show_msg)
                    sys.stdout.flush()
                    if temp_size >= total_size:
                        is_finish = True
                    else:
                        is_finish = False
            except:
                showlog.error('')
    print("\n  ==> 文件已全部下载完成，保存位置:", path_local)
    res_dict = {
        'file_dir': path_local,
        'is_finish': is_finish,
        'temp_size': temp_size
    }
    return res_dict


def safe_download(
        url,
        filename=None,
        suffix_name=None,
        headers=None,
        path="download",
        proxies=None,
        size_limit=None,
        range_start=None,
        range_end=None
):
    while True:
        download_response = download(
            url=url,
            filename=filename,
            suffix_name=suffix_name,
            headers=headers,
            path=path,
            proxies=proxies,
            size_limit=size_limit,
            range_start=range_start,
            range_end=range_end
        )
        if download_response.get('is_finish') is True:
            local_file_dir = download_response.get('file_dir')
            return local_file_dir
        else:
            print(':( 下载中断')
            range_start = download_response.get('temp_size')
            time.sleep(1)
            print('将继续下载（断点续传）...')


def read(
        file,
        postfix: str = None,
        path: str = None,
        json_auto: bool = False,
        read_lines: bool = False
):
    """
    读取文件
    json_auto：json格式自动转换
    """
    if path:
        # 如果指定了路径，就加上路径
        file_dir = f'{path}{path_separator}{file}'
    else:
        # 如果没指定路径，就直接使用文件名
        file_dir = file

    if postfix:
        # 如果指定了后缀名，就加上后缀名
        file_dir = f'{file}.{postfix}'
    else:
        # 如果没指定后缀名，就忽略
        pass

    if read_lines:
        with open(file=file_dir, mode='r', encoding='utf-8') as f:
            content = f.readlines()
    else:
        f = open(file=file_dir, mode='r', encoding='utf-8')
        content = f.read()

    if content:
        if json_auto:
            if isinstance(content, str):
                return json.loads(content)
            else:
                json_content = list()
                for each_line in content:
                    json_content.append(json.loads(each_line))
                return json_content
        else:
            return content
    else:
        return content


def save(
        file,
        content,
        postfix: str = None,
        path: str = None,
        overwrite: bool = True,
        encoding: str = 'utf-8'
):
    """
    保存文件
    """
    if path:
        # 如果指定了路径，就加上路径
        lazypath.make_path(path)
        file_dir = f'{path}{path_separator}{file}'
    else:
        # 如果没指定路径，就直接使用文件名
        file_dir = file

    if postfix:
        # 如果指定了后缀名，就加上后缀名
        file_dir = f'{file}.{postfix}'
    else:
        # 如果没指定后缀名，就忽略
        pass

    if overwrite is True:
        write_mode = "w"  # 覆盖
    else:
        write_mode = "a"  # 追加

    f = open(
        file=file_dir,
        mode=write_mode,
        encoding=encoding
    )
    f.write(content)
    f.close()


def read_(_source_file):
    # >>读取数据【方式：一次性全部读取】------------------------------------------------------------
    data = xlrd.open_workbook(_source_file)  # 打开表
    res = list()
    # table = data.sheets()[0]  # 默认使用第一张表格
    for table in data.sheets():
        nrows = table.nrows  # 获取行数
        ncols = table.ncols  # 获取列数
        for inrows in range(nrows):
            res_temp = dict()
            for incols in range(ncols):
                res_temp.update({table.cell(0, incols).value: table.cell(inrows, incols).value})
            # print(res_temp)
            res.append(res_temp)
    return res


def excel_read(source_file):
    # >>读取数据【方式：一次性全部读取】------------------------------------------------------------
    data = xlrd.open_workbook(source_file)  # 打开表
    res = list()
    # table = data.sheets()[0]  # 默认使用第一张表格
    for table in data.sheets():
        nrows = table.nrows  # 获取行数
        ncols = table.ncols  # 获取列数
        for inrows in range(nrows):
            res_temp = dict()
            for incols in range(ncols):
                res_temp.update({table.cell(0, incols).value: table.cell(inrows, incols).value})
            # print(res_temp)
            res.append(res_temp)
    return res


def excel_read_all_table(source_file):
    try:
        # >>读取数据【方式：一次性全部读取】------------------------------------------------------------
        data = xlrd.open_workbook(source_file)  # 打开表
        res = list()
        # table = data.sheets()[0]  # 默认使用第一张表格
        for table in data.sheets():
            nrows = table.nrows  # 获取行数
            ncols = table.ncols  # 获取列数
            for inrows in range(nrows):
                res_temp = dict()
                for incols in range(ncols):
                    res_temp.update({table.cell(0, incols).value: table.cell(inrows, incols).value})
                # print(res_temp)
                res.append(res_temp)
        return res
    except:
        return


def excel_sheet_dict(source_file):
    try:
        # >>读取数据【方式：一次性全部读取】------------------------------------------------------------
        data = xlrd.open_workbook(source_file)  # 打开表
        res = dict()
        for table in data.sheets():
            table_name = table.name
            nrows = table.nrows  # 获取行数
            ncols = table.ncols  # 获取列数
            res_temp = list()
            for inrows in range(nrows):
                data_temp = dict()
                for incols in range(ncols):
                    data_temp.update({table.cell(0, incols).value: table.cell(inrows, incols).value})
                res_temp.append(data_temp)
            res.update({table_name: res_temp})
        return res
    except:
        return


def excel_specify_dict(source_file, sheet_name, _up_data_col_list=None, _col_dict=None):
    try:
        col_dict = dict()
        if _col_dict is not None:
            temp0 = _col_dict.replace(" ", "").split(",")
            if temp0 is not None and len(temp0) > 0:
                for each in temp0:
                    temp1 = each.split(":")
                    col_dict.update({temp1[1]: temp1[0]})
        up_data_col_list = list()
        if _up_data_col_list is not None:
            up_data_col_list = _up_data_col_list.replace(" ", "").split(",")

        # >>读取数据【方式：一次性全部读取】------------------------------------------------------------
        data = xlrd.open_workbook(source_file)  # 打开表
        res = dict()
        for table in data.sheets():
            table_name = table.name
            if table_name == sheet_name:
                nrows = table.nrows  # 获取行数
                ncols = table.ncols  # 获取列数
                res_temp = list()
                for inrows in range(nrows):
                    data_temp = dict()
                    for incols in range(ncols):
                        # print(table.cell(0, incols).value)
                        col_name_ori = table.cell(0, incols).value
                        if len(col_dict) > 0 and col_dict is not None:
                            col_name = col_dict.get(col_name_ori)
                        else:
                            col_name = col_name_ori

                        if col_name_ori in up_data_col_list:
                            # print(col_name_ori)
                            try:
                                col_value = xlrd.xldate.xldate_as_datetime(table.cell(inrows, incols).value, 0)
                            except Exception as ex:
                                # print(ex)
                                col_value = None
                        else:
                            col_value = table.cell(inrows, incols).value

                        data_temp.update({col_name: col_value})
                    res_temp.append(data_temp)
                res.update({table_name: res_temp})
        return res
    except:
        return


def read_txt(text_name, path=None):
    try:
        if path is None:
            f = open("%s.txt" %text_name, encoding='utf-8')
        else:
            f = open("%s/%s.txt" % (path, text_name), encoding='utf-8')
        res = f.read()
        return res
    except:
        return


def read_file(file_name, suffix_name, path=None):
    try:
        if path is None:
            f = open("\%s.%s" % (file_name, suffix_name))
        else:
            f = open("\%s\%s.%s" % (path, file_name, suffix_name))
        res = f.read()
        return res
    except:
        return


def dir_file_list(file_dir):
    file_list = list()
    for root, dirs, files in os.walk(file_dir):
        # print(root)  # 当前目录路径
        # print(dirs)  # 当前路径下所有子目录
        # print(files)  # 当前路径下所有非目录子文件
        file_list.extend(files)
    return file_list


def get_file_size(file_dir):
    # 获取文件大小
    size_byte = os.path.getsize(file_dir)  # 字节
    size_kb = int(size_byte / 1024)
    size_mb = round(size_byte / (1024 * 1024), 2)
    size_gb = round(size_byte / (1024 * 1024 * 1024), 2)
    if size_gb < 1:
        if size_mb < 1:
            size_str = str(size_kb) + ' KB'
        else:
            size_str = str(size_mb) + ' MB'
    else:
        size_str = str(size_gb) + ' GB'
    temp_dict = {
        'size_str': size_str,
        'size_byte': size_byte
    }
    return temp_dict


class DateEncoder(json.JSONEncoder):
    # 处理json.dumps中会出现的datetime格式无法转换问题：json.dumps(each, cls=DateEncoder)
    def default(self, obj):
        if isinstance(obj, datetime.datetime):
            return obj.strftime("%Y-%m-%d %H:%M:%S")
        else:
            return json.JSONEncoder.default(self, obj)


def save_list(
        file_name,
        list_data,
        split_by='\n',
        postfix="txt",
        path=None,
        overwrite=False
):
    """
    将list按照行存储到文件中，对于datetime类型自动转换为日期时间的格式
    """
    if path is None:
        file_dir = "%s.%s" % (file_name, postfix)
    else:
        lazypath.make_path(path)
        file_dir = "%s%s%s.%s" % (path, path_separator, file_name, postfix)

    if overwrite is True:

        write_mode = "w"  # 覆盖
    else:
        write_mode = "a"  # 追加

    f = open(file_dir, write_mode, encoding='utf-8')
    for each in list_data:

        if isinstance(each, collections.OrderedDict):
            f.write(str(json.dumps(each, cls=DateEncoder)))
        else:
            f.write(str(each))
        f.write(split_by)
    f.close()
    return file_dir


def dict_write07excel(path, table_name, sheet_name, data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name
    if len(data) > 0:
        row_num = 1
        col_num = 1
        for key in data[0]:
            sheet.cell(row=row_num, column=col_num, value=key)
            col_num += 1
        row_num += 1

        for each in data:
            col_num = 1
            for key in each:
                sheet.cell(row=row_num, column=col_num, value=each.get(key))
                col_num += 1
            row_num += 1

        wb.save("%s/%s.xlsx" % (path, table_name))
        showlog.info("导出数据成功！共计%s条数据" % len(data))


def save_txt(
        text_name,
        content,
        path=None,
        overwrite=True):
    """
    是否覆写看overwrite参数
    设置自动创建目录
    """
    if overwrite is True:
        mode = "w"
    else:
        mode = "a"

    if path is None:
        f = open("%s.txt" % text_name, mode, encoding='utf-8')
    else:
        lazypath.make_path(path)
        f = open("%s%s%s.txt" % (path, path_separator, text_name), mode, encoding='utf-8')
    f.write(content)
    f.close()


def save_sql(
        file_name,
        content,
        path=None
):
    """
    如果文件存在，先清空，再保存
    :param file_name:
    :param content:
    :param path:
    :return:
    """
    if path is None:
        f = open(file_name, "w", encoding='utf-8')
    else:
        f = open("%s%s%s.sql" % (path, path_separator, file_name), "w", encoding='utf-8')
    f.write(content)
    f.close()
