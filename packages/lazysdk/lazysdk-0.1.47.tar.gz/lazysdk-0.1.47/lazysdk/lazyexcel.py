#!/usr/bin/env python3
# coding = utf8
"""
@ Author : ZeroSeeker
@ e-mail : zeroseeker@foxmail.com
@ GitHub : https://github.com/ZeroSeeker
@ Gitee : https://gitee.com/ZeroSeeker
"""
import decimal
from collections import OrderedDict
from . import lazydict
import openpyxl  # 用于处理.xlsx，文档：https://openpyxl.readthedocs.io/en/stable/
import datetime
import xlrd  # 用于处理.xls
import os
from .lazypath import path_separator
from .lazytime import get_format_date
from .lazytime import get_format_datetime


# -------------------------------读取功能区域-------------------------------
def read_xls(
        file: str,
        sheet_name: str = None,
        date_cols: dict = None,
        auto_conv_date: bool = False  # 默认格式：'%Y-%m-%d %H:%M:%S'
) -> dict:
    """
    读取xls表格
    :param file: 文件路径
    :param sheet_name: 指定sheet名称，读取指定的sheet
    :param date_cols: 按照既定规则转换时间，例如：{'日期': '%Y-%m-%d','时间': '%H:%M:%S'}，就会将表中对应字段转换为对应格式的字符串
    :param auto_conv_date: 自动转换，会将时间类型的数据自动转换为默认格式的字符串

    :returns: 读取结果字典，字典的第一层key为sheet_name

    cell(a,b).ctype 值类型
    0 empty,
    1 string,
    2 number,
    3 date,
    4 boolean,
    5 error
    """
    if date_cols is None:
        data_cols_cols = []
    else:
        data_cols_cols = date_cols.keys()
    table_data_dict = OrderedDict()  # 有序字典
    table_data = xlrd.open_workbook(file)  # 打开表
    for each_sheet in table_data.sheets():  # 遍历所有sheet
        if sheet_name is None:
            pass
        else:
            if each_sheet.name == sheet_name:  # 判断当前的sheet是否是需要获取数据的sheet
                pass
            else:
                continue
        row_num = each_sheet.nrows  # 获取行数
        col_num = each_sheet.ncols  # 获取列数
        sheet_data_list = list()
        for each_row in range(1, row_num):  # 遍历每一行
            each_row_data = dict()
            for each_col in range(col_num):  # 对于当前行，遍历每一列，将每一行的数据以dict的形式组织到each_row_data里面
                col_name = each_sheet.cell(0, each_col).value  # 列名
                if col_name is None:
                    continue
                else:
                    pass
                col_value = each_sheet.cell(each_row, each_col).value  # 值
                col_type = each_sheet.cell(each_row, each_col).ctype  # 值类型
                if col_value is None:
                    pass
                else:
                    if col_type == 3:
                        if col_name in data_cols_cols:
                            col_value = xlrd.xldate_as_datetime(col_value, 0).strftime(date_cols[col_name])  # 转换为datatime后转换为需要的格式
                        else:
                            if len(data_cols_cols) == 0 and auto_conv_date is True:
                                col_value = xlrd.xldate_as_datetime(col_value, 0).strftime('%Y-%m-%d %H:%M:%S')
                            else:
                                pass
                    else:
                        pass
                each_row_data[col_name] = col_value
            each_row_data_single = list(set(each_row_data.values()))
            if len(each_row_data_single) == 1:
                if each_row_data_single[0] is None:
                    continue  # 去除全空值
                else:
                    pass
            else:
                pass
            sheet_data_list.append(each_row_data)  # 将每一行的数据dict组织到当前sheet的数据list中
        table_data_dict[each_sheet.name] = sheet_data_list  # 将每个sheet的数据组织到table_data_dict里面
    return table_data_dict


def read_xlsx(
        file: str,
        sheet_name: str = None,
        date_cols: dict = None,
        auto_conv_date: bool = False  # 默认格式：'%Y-%m-%d %H:%M:%S'
) -> dict:
    """
    读取xlsx表格
    :param file: 文件路径
    :param sheet_name: 指定sheet名称，读取指定的sheet
    :param date_cols: 按照既定规则转换时间，例如：{'日期': '%Y-%m-%d','时间': '%H:%M:%S'}，就会将表中对应字段转换为对应格式的字符串
    :param auto_conv_date: 自动转换，会将时间类型的数据自动转换为默认格式的字符串

    :returns: 读取结果字典，字典的第一层key为sheet_name

    cell(a,b).data_type 值类型
    n None
    s string
    d date
    其他暂时未知
    """
    if date_cols is None:
        data_cols_cols = []
    else:
        data_cols_cols = date_cols.keys()
    table_data_dict = OrderedDict()  # 有序字典
    table_data = openpyxl.load_workbook(file)  # 打开表
    for each_sheet in table_data.worksheets:  # 遍历所有sheet
        if sheet_name is None:
            pass
        else:
            if each_sheet.title == sheet_name:  # 判断当前的sheet是否是需要获取数据的sheet
                pass
            else:
                continue
        row_num = each_sheet.max_row  # 获取行数
        col_num = each_sheet.max_column  # 获取列数
        sheet_data_list = list()
        for each_row in range(1, row_num):  # 遍历每一行
            each_row_data = dict()
            for each_col in range(1, col_num + 1):  # 对于当前行，遍历每一列，将每一行的数据以dict的形式组织到each_row_data里面
                col_name = each_sheet.cell(1, each_col).value
                if col_name is None:
                    continue
                else:
                    pass
                col_value = each_sheet.cell(each_row + 1, each_col).value  # 原来为日期的会自动转换
                col_type = each_sheet.cell(each_row + 1, each_col).data_type
                if col_value is None:
                    pass
                else:
                    if col_type == 'd':
                        if len(data_cols_cols) == 0 and auto_conv_date is True:
                            col_value = col_value.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            if col_name in data_cols_cols:
                                col_value = col_value.strftime(date_cols[col_name])  # 已经转换的直接按照目标格式转换
                            else:
                                pass
                    else:
                        pass
                each_row_data[col_name] = col_value
            each_row_data_single = list(set(each_row_data.values()))
            if len(each_row_data_single) == 1:
                if each_row_data_single[0] is None:
                    continue  # 去除全空值
                else:
                    pass
            else:
                pass
            sheet_data_list.append(each_row_data)  # 将每一行的数据dict组织到当前sheet的数据list中
        table_data_dict[each_sheet.title] = sheet_data_list  # 将每个sheet的数据组织到table_data_dict里面
    return table_data_dict


def read(
        file: str,
        sheet_name: str = None,
        date_cols: dict = None,  # date转换为字符串
        auto_conv_date: bool = False  # 默认格式：'%Y-%m-%d %H:%M:%S'
) -> dict:
    """
    读取表格数据，支持xls和xlsx两种格式，会自动判断
    :param file: 读取指定文件
    :param sheet_name: 读取指定sheet，一旦指定，将只读取这个sheet的内容
    :param date_cols: 按照既定规则转换时间，例如：{'日期': '%Y-%m-%d','时间': '%H:%M:%S'}，就会将表中对应字段转换为对应格式的字符串
    :param auto_conv_date: 自动转换，会将时间类型的数据自动转换为默认格式的字符串

    :return: 表格中的所有数据，以sheet的名称为key的dict（有序），每个value是由值字典组成的list
    """
    suffix = file.split('.')[-1]  # 获取文件后缀名
    if suffix == 'xlsx':
        table_data_dict = read_xlsx(
            file=file,
            sheet_name=sheet_name,
            date_cols=date_cols,
            auto_conv_date=auto_conv_date
        )
        return table_data_dict
    elif suffix == 'xls':
        table_data_dict = read_xls(
            file=file,
            sheet_name=sheet_name,
            date_cols=date_cols,
            auto_conv_date=auto_conv_date
        )
        return table_data_dict
    else:
        return OrderedDict()


# -------------------------------读取功能区域-------------------------------
# -------------------------------写入功能区域-------------------------------


def save_xlsx(
        file: str,
        value: dict = None,
        date_cols: list = None,
        datetime_cols: list = None,
        num_cols: list = None,
        col_name_dict: dict = None,
        col_name_sort: list = None
):
    """
    如果输入的value是乱序，将重新排序
    :param file: 文件路径
    :param value: 需要保存的数据，按照{"sheet名": [数据列表]}的形式
    :param date_cols: 按照既定规则转换日期，例如：['日期']，就会将表中对应字段转换为对应格式的日期
    :param datetime_cols: 按照既定规则转换时间，例如：['时间']，就会将表中对应字段转换为对应格式的时间
    :param num_cols: 数字列列表
    :param col_name_dict: 自定义列名的对照关系，规则为：{'旧名称1':'新名称1', '旧名称2':'新名称2'}
    :param col_name_sort: 自定义列名排序，将按照列表顺序排，如果不在列表中，将随机
    将输出保存后的文件绝对路径
    """
    if os.path.isabs(file):  # 判断是否为绝对路径
        pass
    else:
        file = os.getcwd() + path_separator + file
    if value is None:
        return
    else:
        wb = openpyxl.Workbook()
        sheet_index = 0  # sheet序号

        if not num_cols:
            num_cols = []
        if not date_cols:
            date_cols = []
        if not datetime_cols:
            datetime_cols = []

        for sheet_name, sheet_data in value.items():
            # 遍历每个要存储的sheet,sheet_name为sheet的名称，sheet_data为对应sheet的数据
            sheet = wb.create_sheet(
                title=sheet_name,
                index=sheet_index
            )  # 创建一个sheet
            row_num = 1  # 行序号
            col_num = 1  # 列序号
            if sheet_data:
                # 先按照要求处理数据，处理好数据后统一排序

                # 如果数据是乱序，需要先排序
                sheet_data_f = lazydict.list_same_order_dict(
                    list_data=sheet_data,
                    keys_sort=col_name_sort
                )  # 先对要存储的数据做排序对齐

                if date_cols is not None:
                    for each_sheet_data_f in sheet_data_f:
                        # 日期格式化
                        for date_key in date_cols:
                            date_value = each_sheet_data_f.get(date_key)
                            if date_value:
                                try:
                                    each_sheet_data_f[date_key] = get_format_date(date_ori=date_value)
                                except:
                                    pass
                            else:
                                continue
                        # 时间格式化
                        for datetime_key in datetime_cols:
                            datetime_value = each_sheet_data_f.get(datetime_key)
                            if datetime_value:
                                try:
                                    each_sheet_data_f[datetime_key] = get_format_datetime(datetime_ori=datetime_value)
                                except:
                                    pass
                            else:
                                continue

                        # 数字格式化
                        for each_num_col in num_cols:
                            date_value = each_sheet_data_f.get(each_num_col)
                            if date_value:
                                try:
                                    each_sheet_data_f[each_num_col] = decimal.Decimal(date_value)
                                except:
                                    pass
                            else:
                                pass
                else:
                    pass
                # 写入标题行
                for key in sheet_data_f[0]:
                    if col_name_dict is not None and col_name_dict.get(key) is not None:
                        sheet.cell(
                            row=row_num,
                            column=col_num,
                            value=col_name_dict.get(key)
                        )  # 写入标题行
                    else:
                        sheet.cell(
                            row=row_num,
                            column=col_num,
                            value=key
                        )  # 写入标题行
                    col_num += 1

                # 写入数据行
                row_num += 1
                for each in sheet_data_f:
                    col_num = 1
                    for key in each:
                        sheet.cell(
                            row=row_num,
                            column=col_num,
                            value=each.get(key)
                        )
                        col_num += 1
                    row_num += 1
            else:
                continue
            sheet_index += 1
        wb.remove(wb.worksheets[-1])  # 删除最后一个默认的sheet
        wb.save(file)
        return file


def save(
        file: str,
        value: dict = None,
        date_cols: dict = None,  # 字符串转换为date
        overwrite: bool = False,
        col_name_dict: dict = None
):
    """
    如果输入的value是乱序，将重新排序
    :param file: 文件路径，支持：xlsx
    :param value: 需要保存的数据
    :param date_cols: 按照既定规则转换时间，例如：{'日期': '%Y-%m-%d','时间': '%H:%M:%S'}，就会将表中对应字段转换为对应格式的时间
    :param overwrite: 是否覆盖
    :param col_name_dict: 自定义列名的对照关系，规则为：{'旧名称1':'新名称1', '旧名称2':'新名称2'}
    """
    if overwrite is False:
        if os.path.exists(file) is True:
            return
        else:
            pass
    else:
        pass

    suffix = file.split('.')[-1]  # 获取文件后缀名
    if suffix == 'xlsx':
        save_xlsx(
            file=file,
            value=value,
            date_cols=date_cols,
            col_name_dict=col_name_dict
        )
        return
    elif suffix == 'xls':
        print(':( xls is not supported')
        return
    else:
        print(':( file is not supported')
        return
