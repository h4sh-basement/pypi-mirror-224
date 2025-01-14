"""
@author: lijc210@163.com
@file: 1.py
@time: 2019/11/27
@desc: 功能描述。
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb = Workbook()

# 获取当前活跃的worksheet,默认就是第一个worksheet
ws = wb.active

# 设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
ws.cell(row=1, column=1).value = 6

# 从第2行开始，写入9行10列数据，值为对应的列序号A、B、C、D...
for row in range(2, 11):
    for col in range(1, 11):
        ws.cell(row=row, column=col).value = get_column_letter(col)

# 可以使用append插入一行数据
ws.append(["我", "你", "她"])
ws.append(["1", "2", "3"])

# 保存
wb.save(filename="a.xlsx")
